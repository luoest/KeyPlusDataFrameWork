import openpyxl
import time
from openpyxl.styles import Font

class Exceling(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = None
        self.RGB_dict = {'red': 'FFFF3030', 'blue': '0000FF'}
        self.now = time.strftime('%Y-%m-%d %H:%M:%S')

    def getWorkbook(self, excelPath):
        try:
            self.workbook = openpyxl.load_workbook(excelPath)
        except Exception as e:
            raise e
        self.excelFile = excelPath
        return self.workbook

    def getSheet(self, sheetName):
        sheet = self.workbook[sheetName]
        return sheet

    def getRowAndColNum(self, sheet):
        rowNum = sheet.max_row
        colNum = sheet.max_column
        return rowNum, colNum

    def getValueByRow(self, sheet, rowNo):
        return list(sheet.rows)[rowNo-1]

    def getValueByCol(self, sheet, colNo):
        return list(sheet.columns)[colNo-1]

    def getCellValue(self, sheet, cellName):
        return sheet[cellName].value

    def getWriteByRowAndCol(self, sheet, content, rowNo, colNo, style = None):
        sheet.cell(row = rowNo, column = colNo).value = content
        if style:
            sheet.cell(row = rowNo, column = colNo).font = Font(color = self.RGB_dict[style])
        self.workbook.save(self.excelFile)

    def getWriteTime(self, sheet, nowTime, rowNo, colNo, style = None):
        self.now = time.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row = rowNo, column = colNo).value = nowTime
        if style:
            sheet.cell(row = rowNo, column = colNo).font = Font(color = self.RGB_dict[style])
        self.workbook.save(self.excelFile)

    def getWriteByCell(self, sheet, content, cellName, style = None):
        sheet[cellName].value = content
        if style:
            sheet[cellName].font = Font(color = self.RGB_dict[style])
        self.workbook.save(self.excelFile)



if __name__ == '__main__':
    from utilities.getFilePath import excel_path
    from performance.PagePerformance import getOpenLocalFile
    excelObj = Exceling()
    excelObj.getWorkbook(excel_path)
    sheet = excelObj.getSheet('登录')
    print(excelObj.getRowAndColNum(sheet))
    rowValues = excelObj.getValueByRow(sheet, 3)
    for rv in rowValues:
        print(rv.value, end=', ')
    print()
    for colValue in excelObj.getValueByCol(sheet, 3):
        print(colValue.value, end = ',')
    print()
    print('"C3的值是："', excelObj.getCellValue(sheet, 'C3'))

    excelObj.getWriteByRowAndCol(sheet, '今天5月4日，星期六', 2, 1, 'blue')
    excelObj.getWriteTime(sheet, excelObj.now, 3, 1, 'red')
    excelObj.getWriteByCell(sheet, '写入A4', 'A4', 'blue')

    getOpenLocalFile(excel_path)
